from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


WORKBOOK_PATH = Path(__file__).resolve().parents[2] / "virtual_players_2008_complete_with_all_staff_data.xlsx"


ENUM_DEFINITIONS: dict[str, dict[str, str]] = {
    "position": {
        "GK": "Goalkeeper",
        "CB": "Center back",
        "LB": "Left back",
        "RB": "Right back",
        "DM": "Defensive midfielder",
        "CM": "Central midfielder",
        "AM": "Attacking midfielder",
        "LW": "Left winger",
        "RW": "Right winger",
        "CF": "Center forward",
        "ST": "Striker",
        "LM": "Left midfielder",
        "RM": "Right midfielder",
    },
    "dominant_foot": {
        "right": "Right-footed",
        "left": "Left-footed",
        "both": "Two-footed",
    },
    "player_status": {
        "active": "Available for selection",
        "injured": "Unavailable due to injury",
    },
    "injury_severity": {
        "minor": "Short-term issue",
        "moderate": "Medium-term issue",
        "severe": "Long-term issue",
    },
    "injury_status": {
        "rehab": "In rehabilitation",
        "recovered": "Returned to play",
    },
    "boolean_flag": {
        "true": "Yes / true",
        "false": "No / false",
    },
    "injury_context": {
        "match": "Occurred during a match",
        "training": "Occurred during training",
        "outside": "Occurred outside team activity",
    },
    "match_type": {
        "공식": "Official match",
        "연습": "Friendly / practice match",
    },
    "match_phase": {
        "동계훈련 연습경기": "Winter camp friendly",
        "2월 공식대회": "February official competition",
        "주말리그": "Weekend league",
        "주중 연습경기": "Midweek friendly",
        "5월 공식대회": "May official competition",
        "7월 전국대회": "July national tournament",
        "10월 전국대회": "October national tournament",
        "시즌 마무리 연습경기": "Season closing friendly",
    },
    "goal_type": {
        "header": "Header goal",
        "inside_box": "Goal from inside the box",
        "penalty": "Penalty goal",
    },
    "training_type": {
        "conditioning": "Conditioning session",
        "pre_match": "Pre-match session",
        "recovery": "Recovery session",
        "tactical": "Tactical session",
        "tactical_physical": "Integrated tactical and physical session",
        "technical": "Technical session",
    },
    "training_focus": {
        "고강도 전술 + 체력": "High-intensity tactics and conditioning",
        "기술 완성도": "Technical refinement",
        "전술 정리 및 세트피스": "Tactical organization and set-pieces",
        "조직 전술": "Team tactical organization",
        "체력·파워 향상": "Conditioning and power development",
        "회복 및 재생": "Recovery and regeneration",
    },
    "session_name": {
        "경기 전날 프리매치 세션": "Pre-match session",
        "경기 후 회복훈련": "Post-match recovery",
        "기술 완성 훈련": "Technical finishing session",
        "동계 고강도 피지컬 세션": "Winter high-intensity physical session",
        "전술 조직훈련": "Tactical organization session",
        "주중 고강도 전술훈련": "Midweek high-intensity tactical session",
    },
    "intensity_level": {
        "low": "Low intensity",
        "medium": "Medium intensity",
        "high": "High intensity",
    },
    "counseling_topic": {
        "경기 피드백": "Match feedback",
        "멘탈 관리": "Mental management",
        "부상 관리": "Injury management",
        "진로 상담": "Career counseling",
        "훈련 태도": "Training attitude",
    },
}


ENUM_ALIASES: dict[str, dict[Any, str]] = {
    "position": {
        "gk": "GK",
        "cb": "CB",
        "lb": "LB",
        "rb": "RB",
        "dm": "DM",
        "cm": "CM",
        "am": "AM",
        "lw": "LW",
        "rw": "RW",
        "cf": "CF",
        "st": "ST",
        "lm": "LM",
        "rm": "RM",
    },
    "dominant_foot": {
        "right": "right",
        "left": "left",
        "both": "both",
        "two-footed": "both",
    },
    "player_status": {
        "active": "active",
        "available": "active",
        "injured": "injured",
    },
    "injury_severity": {
        "minor": "minor",
        "moderate": "moderate",
        "severe": "severe",
    },
    "injury_status": {
        "rehab": "rehab",
        "recovered": "recovered",
        "returned": "recovered",
    },
    "boolean_flag": {
        True: "true",
        False: "false",
        1: "true",
        0: "false",
        "true": "true",
        "false": "false",
        "yes": "true",
        "no": "false",
        "y": "true",
        "n": "false",
    },
    "injury_context": {
        "match": "match",
        "training": "training",
        "outside": "outside",
        "external": "outside",
    },
    "match_type": {
        "공식": "공식",
        "연습": "연습",
        "official": "공식",
        "friendly": "연습",
        "practice": "연습",
    },
    "match_phase": {
        value: value for value in ENUM_DEFINITIONS["match_phase"]
    },
    "goal_type": {
        "header": "header",
        "inside_box": "inside_box",
        "penalty": "penalty",
    },
    "training_type": {
        value: value for value in ENUM_DEFINITIONS["training_type"]
    },
    "training_focus": {
        value: value for value in ENUM_DEFINITIONS["training_focus"]
    },
    "session_name": {
        value: value for value in ENUM_DEFINITIONS["session_name"]
    },
    "intensity_level": {
        "low": "low",
        "medium": "medium",
        "high": "high",
    },
    "counseling_topic": {
        value: value for value in ENUM_DEFINITIONS["counseling_topic"]
    },
}


@dataclass(frozen=True)
class EnumBinding:
    sheet_name: str
    column_name: str
    enum_key: str
    allow_blank: bool = False


ENUM_BINDINGS: tuple[EnumBinding, ...] = (
    EnumBinding("player_info", "primary_position", "position"),
    EnumBinding("player_info", "secondary_position", "position", allow_blank=True),
    EnumBinding("player_info", "foot", "dominant_foot"),
    EnumBinding("player_info", "status", "player_status"),
    EnumBinding("injury_history", "severity_level", "injury_severity"),
    EnumBinding("injury_history", "status", "injury_status"),
    EnumBinding("injury_history", "surgery_required", "boolean_flag"),
    EnumBinding("injury_history", "occurred_during", "injury_context"),
    EnumBinding("match_data", "match_type", "match_type"),
    EnumBinding("match_data", "phase", "match_phase"),
    EnumBinding("match_player_data", "position", "position"),
    EnumBinding("match_player_data", "start_position", "position", allow_blank=True),
    EnumBinding("match_player_data", "goals_type", "goal_type", allow_blank=True),
    EnumBinding("training_data", "training_type", "training_type"),
    EnumBinding("training_data", "training_focus", "training_focus"),
    EnumBinding("training_data", "session_name", "session_name"),
    EnumBinding("training_data", "intensity_level", "intensity_level"),
    EnumBinding("counseling", "topic", "counseling_topic"),
)


def canonicalize_enum_value(enum_key: str, value: Any) -> str | None:
    if value is None:
        return None
    if value != value:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        raw: Any = stripped
    else:
        raw = value

    if raw in ENUM_DEFINITIONS[enum_key]:
        return raw  # type: ignore[return-value]

    alias_lookup = ENUM_ALIASES.get(enum_key, {})
    if raw in alias_lookup:
        return alias_lookup[raw]

    if isinstance(raw, str):
        lowered = raw.casefold()
        if lowered in alias_lookup:
            return alias_lookup[lowered]

    raise ValueError(f"Unsupported value for {enum_key}: {value!r}")


def apply_enum_reference(path: Path = WORKBOOK_PATH) -> None:
    workbook = load_workbook(path)

    for binding in ENUM_BINDINGS:
        if binding.sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[binding.sheet_name]
        headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        if binding.column_name not in headers:
            continue
        column_idx = headers[binding.column_name]
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=column_idx)
            cell.value = canonicalize_enum_value(binding.enum_key, cell.value)

    if "player_info" in workbook.sheetnames:
        ws = workbook["player_info"]
        headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        primary_idx = headers.get("primary_position")
        secondary_idx = headers.get("secondary_position")
        if primary_idx and secondary_idx:
            for row_idx in range(2, ws.max_row + 1):
                primary = ws.cell(row=row_idx, column=primary_idx).value
                secondary = ws.cell(row=row_idx, column=secondary_idx).value
                if primary and secondary and primary == secondary:
                    ws.cell(row=row_idx, column=secondary_idx).value = None

    if "enum_reference" in workbook.sheetnames:
        workbook.remove(workbook["enum_reference"])
    enum_sheet = workbook.create_sheet("enum_reference")

    for enum_idx, (enum_key, values) in enumerate(ENUM_DEFINITIONS.items(), start=1):
        code_col = (enum_idx - 1) * 2 + 1
        label_col = code_col + 1
        enum_sheet.cell(row=1, column=code_col, value=enum_key)
        enum_sheet.cell(row=1, column=label_col, value=f"{enum_key}__label")
        enum_sheet.cell(row=2, column=code_col, value="code")
        enum_sheet.cell(row=2, column=label_col, value="description")

        for row_offset, (code, description) in enumerate(values.items(), start=3):
            enum_sheet.cell(row=row_offset, column=code_col, value=code)
            enum_sheet.cell(row=row_offset, column=label_col, value=description)

        range_end = len(values) + 2
        range_ref = f"'enum_reference'!${get_column_letter(code_col)}$3:${get_column_letter(code_col)}${range_end}"
        name = f"{enum_key}_enum"
        if name in workbook.defined_names:
            del workbook.defined_names[name]
        workbook.defined_names.add(DefinedName(name=name, attr_text=range_ref))

    for binding in ENUM_BINDINGS:
        if binding.sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[binding.sheet_name]
        headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        if binding.column_name not in headers:
            continue
        column_letter = get_column_letter(headers[binding.column_name])
        ws.data_validations.dataValidation = [
            dv
            for dv in ws.data_validations.dataValidation
            if f"{column_letter}2" not in str(dv.sqref)
        ]
        validation = DataValidation(
            type="list",
            formula1=f"={binding.enum_key}_enum",
            allow_blank=binding.allow_blank,
        )
        validation.prompt = f"Allowed values: {', '.join(ENUM_DEFINITIONS[binding.enum_key].keys())}"
        validation.error = f"{binding.column_name} must use the defined enum list."
        ws.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}1048576")

    workbook.save(path)
