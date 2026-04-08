from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .calendar_update_shared import (
    COUNSELING_COLUMNS,
    DATETIME_COLUMNS,
    DATE_ONLY_COLUMNS,
    EVALUATION_COLUMNS,
    GPS_LEADING_COLUMNS,
    MATCH_GPS_LEADING_COLUMNS,
    MATCH_PLAYER_CONTEXT_COLUMNS,
    TRAINING_GPS_LEADING_COLUMNS,
    WorkbookFrames,
    build_year_scoped_ids,
    normalize_date,
    normalize_datetime,
)


def select_columns(frame: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    return frame.loc[:, list(columns)].copy()


def compose_output_columns(
    frame: pd.DataFrame,
    always_columns: Sequence[str],
    source_columns: Sequence[str],
    *,
    optional_columns: Sequence[str] = (),
) -> list[str]:
    ordered_columns = [column for column in always_columns if column in frame.columns]
    for column in optional_columns:
        if column in source_columns and column in frame.columns and column not in ordered_columns:
            ordered_columns.append(column)
    for column in source_columns:
        if column in frame.columns and column not in ordered_columns:
            ordered_columns.append(column)
    return ordered_columns


def build_gps_output_frames(
    gps_frame: pd.DataFrame,
    match_data: pd.DataFrame,
    training_data: pd.DataFrame,
    gps_sheet_mode: str,
    source_columns: dict[str, list[str]],
) -> WorkbookFrames:
    match_date_lookup = match_data.set_index("match_id")["match_date"].to_dict()
    training_date_lookup = training_data.set_index("training_id")["training_date"].to_dict()

    if gps_sheet_mode == "combined":
        gps = gps_frame.copy()
        gps["event_date"] = gps["match_id"].map(match_date_lookup)
        gps["event_date"] = gps["event_date"].fillna(gps["training_id"].map(training_date_lookup))
        gps["gps_id"] = build_year_scoped_ids(gps, "event_date", "GPS", 5)
        gps = gps.drop(columns=["event_date"])
        return {
            "gps_data": select_columns(
                gps,
                compose_output_columns(gps, GPS_LEADING_COLUMNS, source_columns["gps_data"]),
            )
        }

    match_gps = gps_frame.loc[gps_frame["match_id"].notna()].copy()
    match_gps["event_date"] = match_gps["match_id"].map(match_date_lookup)
    match_gps["match_gps_id"] = build_year_scoped_ids(match_gps, "event_date", "MGPS", 5)
    match_gps = match_gps.drop(columns=["training_id", "event_date"], errors="ignore")

    training_gps = gps_frame.loc[gps_frame["training_id"].notna()].copy()
    training_gps["event_date"] = training_gps["training_id"].map(training_date_lookup)
    training_gps["training_gps_id"] = build_year_scoped_ids(training_gps, "event_date", "TGPS", 5)
    training_gps = training_gps.drop(columns=["match_id", "event_date"], errors="ignore")

    return {
        "match_gps_data": select_columns(
            match_gps,
            compose_output_columns(match_gps, MATCH_GPS_LEADING_COLUMNS, source_columns["match_gps_data"]),
        ),
        "training_gps_data": select_columns(
            training_gps,
            compose_output_columns(training_gps, TRAINING_GPS_LEADING_COLUMNS, source_columns["training_gps_data"]),
        ),
    }


def build_output_frames(
    frames: WorkbookFrames,
    gps_sheet_mode: str,
    source_columns: dict[str, list[str]],
) -> WorkbookFrames:
    match_player_optional_columns = (
        *MATCH_PLAYER_CONTEXT_COLUMNS,
        "goals_for",
        "goals_against",
        "possession_for",
        "possession_against",
        "player_name",
        "player_birth_day",
    )
    output_frames: WorkbookFrames = {
        "player_info": select_columns(frames["player_info"], source_columns["player_info"]),
        "physical_test_data": select_columns(
            frames["physical_test_data"].drop(columns=["year"], errors="ignore"),
            compose_output_columns(
                frames["physical_test_data"].drop(columns=["year"], errors="ignore"),
                ("physical_test_id", "player_id", "test_date"),
                source_columns["physical_test_data"],
                optional_columns=("player_name", "player_birth_day"),
            ),
        ),
        "physical_data": select_columns(
            frames["physical_data"].drop(columns=["year"], errors="ignore"),
            compose_output_columns(
                frames["physical_data"].drop(columns=["year"], errors="ignore"),
                ("physical_data_id", "player_id", "height", "weight", "body_fat_percentage", "bmi", "muscle_mass", "created_at"),
                source_columns["physical_data"],
                optional_columns=("player_name", "player_birth_day"),
            ),
        ),
        "injury_history": select_columns(frames["injury_history"], source_columns["injury_history"]),
        "match_data": select_columns(
            frames["match_data"].drop(columns=["home_team", "away_team", "year"], errors="ignore"),
            compose_output_columns(
                frames["match_data"].drop(columns=["home_team", "away_team", "year"], errors="ignore"),
                ("match_id", "match_date", "match_type", "phase", "stadium", "opponent_team"),
                source_columns["match_data"],
            ),
        ),
        "match_player_data": select_columns(
            frames["match_player_data"].drop(columns=["home_team", "away_team"], errors="ignore"),
            compose_output_columns(
                frames["match_player_data"].drop(columns=["home_team", "away_team"], errors="ignore"),
                ("match_player_id", "match_id", "player_id", "position", "minutes_played", "start_position", "substitute_in", "substitute_out"),
                source_columns["match_player_data"],
                optional_columns=match_player_optional_columns,
            ),
        ),
        "training_data": select_columns(
            frames["training_data"].drop(columns=["year"], errors="ignore"),
            source_columns["training_data"],
        ),
        "evaluations": select_columns(
            frames["evaluations"],
            compose_output_columns(frames["evaluations"], EVALUATION_COLUMNS, source_columns["evaluations"]),
        ),
        "counseling": select_columns(
            frames["counseling"],
            compose_output_columns(frames["counseling"], COUNSELING_COLUMNS, source_columns["counseling"]),
        ),
    }
    output_frames.update(
        build_gps_output_frames(
            gps_frame=frames["gps_data"],
            match_data=frames["match_data"],
            training_data=frames["training_data"],
            gps_sheet_mode=gps_sheet_mode,
            source_columns=source_columns,
        )
    )
    return output_frames


def write_frames(workbook_path: Path, frames: WorkbookFrames) -> None:
    workbook = load_workbook(workbook_path)
    for sheet_name in ("gps_data", "match_gps_data", "training_gps_data"):
        if sheet_name in workbook.sheetnames and sheet_name not in frames:
            workbook.remove(workbook[sheet_name])
    workbook.save(workbook_path)

    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)


def apply_temporal_formats(workbook_path: Path, sheet_names: Sequence[str]) -> None:
    workbook = load_workbook(workbook_path)
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        headers = {cell.value: index for index, cell in enumerate(worksheet[1], start=1)} if worksheet.max_row else {}

        for column in DATE_ONLY_COLUMNS.get(sheet_name, set()):
            if column not in headers:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=headers[column])
                cell.value = normalize_date(cell.value)
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"

        for column in DATETIME_COLUMNS.get(sheet_name, set()):
            if column not in headers:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=headers[column])
                cell.value = normalize_datetime(cell.value)
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd hh:mm"

    workbook.save(workbook_path)
