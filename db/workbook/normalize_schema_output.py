from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .normalize_schema_shared import (
    DATETIME_COLUMNS,
    DATE_ONLY_COLUMNS,
    DESIRED_SHEET_ORDER,
    REPLACED_SHEETS,
    WorkbookFrames,
    normalize_date,
    normalize_datetime,
)


def replace_sheets(workbook_path: Path, frames: WorkbookFrames) -> None:
    workbook = load_workbook(workbook_path)
    for sheet_name in REPLACED_SHEETS:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
    workbook.save(workbook_path)

    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)


def apply_temporal_formats(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        headers = {cell.value: index for index, cell in enumerate(worksheet[1], start=1)} if worksheet.max_row else {}

        for column in DATE_ONLY_COLUMNS.get(sheet_name, set()):
            if column not in headers:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=headers[column])
                cell.value = normalize_date(cell.value)
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"

        for column in DATETIME_COLUMNS.get(sheet_name, set()):
            if column not in headers:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=headers[column])
                cell.value = normalize_datetime(cell.value)
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd hh:mm"

    workbook.save(workbook_path)


def ensure_readme_sheet(workbook_path: Path, readme: pd.DataFrame | None) -> None:
    if readme is None:
        return
    workbook = load_workbook(workbook_path)
    if "README" in workbook.sheetnames:
        workbook.close()
        return
    workbook.close()
    with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a") as writer:
        readme.to_excel(writer, sheet_name="README", index=False, header=False)


def reorder_workbook_sheets(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    existing = [sheet_name for sheet_name in DESIRED_SHEET_ORDER if sheet_name in workbook.sheetnames]
    extras = [sheet_name for sheet_name in workbook.sheetnames if sheet_name not in existing]
    workbook._sheets = [workbook[sheet_name] for sheet_name in existing + extras]
    workbook.save(workbook_path)
